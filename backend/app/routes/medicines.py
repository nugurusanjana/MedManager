from datetime import datetime
from flask import Blueprint, request, jsonify
from app import db
from app.models import Medicine, Notification
from app.auth_utils import login_required

bp = Blueprint('medicines', __name__)

def ensure_notifications(user):
    """Create notifications for expiring (6 days) and low stock (<=5) medicines."""
    for m in user.medicines:
        if m.is_expiring_soon(6):
            existing = Notification.query.filter_by(user_id=user.id, medicine_id=m.id, type='expiry').first()
            if not existing or (existing.created_at.date() != datetime.utcnow().date()):
                if not existing:
                    n = Notification(user_id=user.id, medicine_id=m.id, type='expiry',
                                    message=f'{m.name} expires on {m.expiry_date}. Replenish soon.')
                    db.session.add(n)
        if m.is_low_stock(5):
            existing = Notification.query.filter_by(user_id=user.id, medicine_id=m.id, type='low_stock').first()
            if not existing or (existing.created_at.date() != datetime.utcnow().date()):
                if not existing:
                    n = Notification(user_id=user.id, medicine_id=m.id, type='low_stock',
                                    message=f'{m.name} is low on stock (quantity: {m.quantity}). Replenish.')
                    db.session.add(n)
    db.session.commit()

@bp.route('/', methods=['GET'])
@login_required
def list_medicines(user):
    ensure_notifications(user)
    q = request.args.get('q', '').strip()
    medicines = Medicine.query.filter_by(user_id=user.id).order_by(Medicine.name).all()
    if q:
        ql = q.lower()
        medicines = [m for m in medicines if ql in (m.name or '').lower() or ql in (m.usage or '').lower()]
    return jsonify([m.to_dict() for m in medicines])

@bp.route('/<int:mid>', methods=['GET'])
@login_required
def get_medicine(user, mid):
    m = Medicine.query.filter_by(id=mid, user_id=user.id).first()
    if not m:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(m.to_dict())

@bp.route('/', methods=['POST'])
@login_required
def add_medicine(user):
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    expiry = data.get('expiry_date')
    quantity = int(data.get('quantity') or 0)
    usage = (data.get('usage') or '').strip()
    cost = float(data.get('cost_per_unit') or 0)
    if not name or not expiry:
        return jsonify({'error': 'name and expiry_date required'}), 400
    try:
        expiry_date = datetime.strptime(expiry[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid expiry_date'}), 400
    m = Medicine(user_id=user.id, name=name, expiry_date=expiry_date, quantity=quantity, usage=usage, cost_per_unit=cost)
    db.session.add(m)
    db.session.commit()
    ensure_notifications(user)
    return jsonify(m.to_dict()), 201

@bp.route('/<int:mid>', methods=['PUT'])
@login_required
def update_medicine(user, mid):
    m = Medicine.query.filter_by(id=mid, user_id=user.id).first()
    if not m:
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json() or {}
    if 'name' in data:
        m.name = (data['name'] or '').strip() or m.name
    if 'expiry_date' in data:
        try:
            m.expiry_date = datetime.strptime(str(data['expiry_date'])[:10], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
    if 'quantity' in data:
        m.quantity = int(data['quantity'])
    if 'usage' in data:
        m.usage = (data['usage'] or '').strip()
    if 'cost_per_unit' in data:
        m.cost_per_unit = float(data['cost_per_unit'])
    db.session.commit()
    ensure_notifications(user)
    return jsonify(m.to_dict())

@bp.route('/<int:mid>', methods=['DELETE'])
@login_required
def delete_medicine(user, mid):
    m = Medicine.query.filter_by(id=mid, user_id=user.id).first()
    if not m:
        return jsonify({'error': 'Not found'}), 404
    
    from app.models import Purchase, CartItem
    Notification.query.filter_by(medicine_id=mid).delete()
    Purchase.query.filter_by(medicine_id=mid).delete()
    CartItem.query.filter_by(medicine_id=mid).delete()
    
    db.session.delete(m)
    db.session.commit()
    return jsonify({'ok': True}), 200

@bp.route('/<int:mid>/quantity', methods=['PATCH'])
@login_required
def adjust_quantity(user, mid):
    data = request.get_json() or {}
    delta = int(data.get('delta', 0))
    m = Medicine.query.filter_by(id=mid, user_id=user.id).first()
    if not m:
        return jsonify({'error': 'Not found'}), 404
    m.quantity = max(0, m.quantity + delta)
    db.session.commit()
    ensure_notifications(user)
    return jsonify(m.to_dict())

def _parse_expiry(val):
    if not val:
        return datetime.utcnow().date()
    if hasattr(val, 'date'):
        return val.date()
    s = str(val).strip()[:10]
    if not s:
        return datetime.utcnow().date()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d-%b-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return datetime.utcnow().date()

def _normalize_key(k):
    """Lowercase, strip spaces/underscores, remove special chars like (₹)."""
    import re
    k = k.lower()
    k = re.sub(r'[^\w\s]', '', k)   # remove special chars
    k = re.sub(r'[\s_]+', '_', k)   # spaces/underscores -> single underscore
    return k.strip('_')

def _get(row, *keys):
    """Fuzzy key lookup — matches regardless of spacing, case, or special chars."""
    normalized_row = {_normalize_key(k): v for k, v in row.items()}
    for k in keys:
        lookup = _normalize_key(k)
        if lookup in normalized_row and normalized_row[lookup] is not None:
            return normalized_row[lookup]
    return None

@bp.route('/bulk', methods=['POST'])
@login_required
def bulk_upload(user):
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    try:
        if f.filename.lower().endswith('.csv'):
            import csv, io
            stream = io.StringIO(f.stream.read().decode('utf-8-sig'))
            reader = csv.DictReader(stream)
            rows = list(reader)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True)
            ws = wb.active
            # preserve original header casing — _get() will normalize at lookup time
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2):
                rows.append(dict(zip(headers, [c.value for c in row])))
            wb.close()
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    created = 0
    for row in rows:
        name = str(_get(row, 'name', 'medicine_name', 'medicine name') or '').strip()
        if not name:
            continue
        expiry  = _get(row, 'expiry_date', 'expiry date', 'expiry', 'exp_date', 'exp date')
        quantity = int(_get(row, 'quantity', 'qty') or 0)
        usage   = str(_get(row, 'usage', 'indication', 'usage / indication') or '').strip()
        cost    = float(_get(row, 'cost_per_unit', 'cost per unit', 'cost', 'price') or 0)

        m = Medicine(
            user_id=user.id,
            name=name,
            expiry_date=_parse_expiry(expiry),
            quantity=quantity,
            usage=usage,
            cost_per_unit=cost
        )
        db.session.add(m)
        created += 1

    db.session.commit()
    ensure_notifications(user)
    return jsonify({'created': created}), 201